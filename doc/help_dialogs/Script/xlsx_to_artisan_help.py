# ABOUT
# Script to take an Excel input file and generate Artisan help dialog code.

# LICENSE
# This program or module is free software: you can redistribute it and/or
# modify it under the terms of the GNU General Public License as published
# by the Free Software Foundation, either version 2 of the License, or
# version 3 of the License, or (at your option) any later version. It is
# provided for educational purposes and is distributed in the hope that
# it will be useful, but WITHOUT ANY WARRANTY; without even the implied
# warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See
# the GNU General Public License for more details.

# AUTHOR
# Dave Baxter, 2023

'''Command line:  xlsx_to_artisan_help.py <filename>|all
**Note that filename should have no suffix, file name only.

The input file is drawn from ../input_files/<filename>.xlsx
'all' will convert every .xlsx file in the ../input_files folder

Output .py files are written to the ../../../src/help folder, aka the artisan/src folder
Output .html files are written to the ../Output_html folder, for inspection or external linking

Note: This program has no error checking or trapping!  TBA.

Excel file format:
    One Excel file for each help dialog.
    One or more tabs, each tab is a separate table within the same help dialog.
    The tab name is not important but it will be used to generate the associated PrettyTable table name.

    Each tab must have the following cells, starting in the upper left cell A1
    Title of the table               [First row, required]
    Top Note                         [Multiple contiguous rows, optional]
    Table headers, aka column titles [One row, optional but if provided must include table data rows]
    Table data rows                  [Multiple contiguous rows, optional but must follow the table headers]
    Bottom Note                      [Multiple contiguous rows, optional]

    QApplication.translate() is not applied to cells formatted as Italic, formatted with any font color
    or if the cell has a numeric value.

    Non-alphanumeric characters must be replaced with their Unicode encoded equivalent. Example, the degree
    symbol º must be entered as \u00b0 in Excel.  Non-breaking space is entered as \u00A0.

    A single quote at the start of a cell string acts as a flag to Excel that the cell contains text. If the
    single quote at the start of a string is to be displayed and passed along to the help code it should
    be escaped (preceded) with another single quote.  Example, to display the string 'Hello' put ''Hello'
    into the cell. Single quotes in the middle or end of a string do not need to be escaped.

    Single blank rows are legal. Consecutive blank rows are interpreted as the end of a table. All cells in
    a blank row must be empty.

    Top Notes: Can be one or more contiguous rows that will be combined in a single block.  Each row starts a new line.
               New lines can be embedded into the notes text.  '\n' without the quotes.
               Each row must begin with one of 'topnote:' or 'tn:' without the quotes, case insensitive.

    Bottom Notes: Can be one or more contiguous rows that will be combined in a single block.  Each row starts a new line.
                  New lines can be embedded into the notes text.  '\n' without the quotes.
                  Each row must begin with one of 'botnote:' or 'bn:' or 'bottomnote:' without the quotes, case insensitive.
    There is no error checking for notes position so if top and bottom notes are reversed bad things will happen.

Known problems:
- openpyxl does not always report the correct last row or column. This will be seen when rendering the HTML.  This is reported
as not fixable. If this happens, go back to the source Excel file and manually delete the excess rows or columns manually.
https://stackoverflow.com/questions/46569496/openpyxl-max-row-and-max-column-wrongly-reports-a-larger-figure#46571480
A hack was added to correct for openpyxl reporting too many rows. There is not a hack for columns at this time.

Recent changes:
- Single quotes now tolerated
- hacked around the problem: The Unicode ellipsis \x85 "…" is not properly handled.  These seem to get into the Excel file when cut and paste from the blog.
Must be replaced with three periods "..." in the Excel file.
- Alt-Enter in Excel to create newlines now tolerated
- adds support for PyQt6

'''

from os.path import dirname, abspath, split, splitext
from os import listdir
import importlib
import re
from time import sleep
import sys
sys.dont_write_bytecode = True  #prevents __pycache__ folder written to help/

from typing import List, Tuple
from openpyxl.worksheet.worksheet import Worksheet # pylint: disable=unused-import

try:
    from PyQt6.QtWidgets import QApplication  # pylint: disable=unused-import
except ImportError:
    from PyQt5.QtWidgets import QApplication # type: ignore # noqa: F401  # pylint: disable=unused-import
from openpyxl import load_workbook


ind = '    '         #indent
nlind = '\n' + ind   #new line plus indent

def translateStr(in_str:str, group:str='HelpDlg') -> str:
    return "QApplication.translate('" + group + "','" + str(in_str) + "')"

def generateRows(ws:Worksheet) -> List[List[str]]:
    all_rows = []
    for row in ws.iter_rows():
        this_row = []
        for cell in row:
            # insert a nonbreaking space into blank cells to keep the row height homgenous
            if cell.value in {None, ''}:
                cell_str = "'&#160;'"
            else:
                if cell.data_type == 's':
                    cell_value = re.subn(r"'",r'&#39;',str(cell.value))  #protect downstream by changing single quotes to double
                    cell_value = re.subn(r'\n',r'\\n',cell_value[0])
                    # do not translate if the cell has italic format or any font color
                    if cell.font.italic or \
                            (cell.font.color is not None and \
                             isinstance(cell.font.color.rgb,str) and \
                             cell.font.color.rgb != 'FF000000'):
                        cell_str = "'" + cell_value[0] + "'"
                    else:
                        cell_str = translateStr(cell_value[0])
                else:
                    # numeric value
                    cell_str = str(cell.value)
                # hack to solve error when elypsis is in the input.
                # Caution that the elypsis character in the line below does not get changed
                repl = re.subn(r'…',r'...',cell_str)
                cell_str = repl[0]

            this_row.append(cell_str)
        all_rows.append(this_row)
    return all_rows

def getTitle(all_rows:List[List[str]],_:Worksheet,nsheet:int) -> str:
    del _
    if nsheet == 0:
        title = nlind + "strlist.append('<b>')"
    else:
        title = nlind + "strlist.append('<br/><br/><b>')"
    title +=  nlind + 'strlist.append(' + str(all_rows[0][0]) + ')'
    title +=  nlind + "strlist.append('</b>')"
    return title

def getNotes(all_rows:List[List[str]],nrows:int,tbl_name:str,notetype:str='top') -> Tuple[str,int]:
    tbl_name = tbl_name + notetype
    tbl_notes = ''
    notes = []
    notes_len = 0
    for idx in range(1, nrows):
        if notetype == 'top':
            gotnote = re.subn(r'topnote:|tn:','',str(all_rows[idx][0]),count=0,flags=re.IGNORECASE)
        else:
            gotnote = re.subn(r'botnote:|bn:|bottomnote:','',str(all_rows[idx][0]),count=0,flags=re.IGNORECASE)
        if gotnote[1]:
            notes.append(gotnote[0])
            notes_len += 1
    tbl_notes += nlind + tbl_name + ' = prettytable.PrettyTable()'
    tbl_notes += nlind + tbl_name + '.header = False'
    tbl_notes += nlind + tbl_name + '.add_row([' + '+newline+'.join(notes) + '])'

    return tbl_notes, notes_len

def getFieldnames(rows:List[str],tbl_name:str) -> str:
    this_row = ','.join(rows)
    return str(tbl_name) + '.field_names = [' + str(this_row) + ']'

def getAddrows(all_rows:List[List[str]],tbl_name:str) -> str:
    addrows = ''
    for idx, row in enumerate(all_rows):
        this_row = ','.join(row)
        if idx > 0:
            addrows += nlind
        addrows += str(tbl_name) + '.add_row([' + this_row + '])'
    return addrows

def buildpyCode(filename_in:str) -> str:
    data_table_attributes = "'width':'100%','border':'1','padding':'1','border-collapse':'collapse'"
    note_table_attributes = "'width':'100%','border':'1','padding':'1','border-collapse':'collapse'"

    outstr = ''

    outstr += '# This is an autogenerated file -- Do not edit'
    outstr += '\n' + '# Edit the source file in artisan/doc/help_dialogs/Input_files'
    outstr += '\n' + '# then execute artisan/doc/help_dialogs/Script/xlsx_to_artisan_help.py'

    # wrap the output with python code to allow it to execute
    outstr += '\n' + 'import prettytable'
    outstr += '\n' + 'import re'
    outstr += '\n' + 'try:'
    outstr += '\n' + '    from PyQt6.QtWidgets import QApplication # @Reimport @UnresolvedImport @UnusedImport # pylint: disable=import-error'
    outstr += '\n' + 'except Exception: # pylint: disable=broad-except'
    outstr += '\n' + '    from PyQt5.QtWidgets import QApplication # type: ignore # @Reimport @UnresolvedImport @UnusedImport'
    outstr += '\n'
    outstr += '\ndef content() -> str:'
    outstr += nlind + 'strlist = []'
    outstr += nlind + "helpstr = ''  # noqa: F841 #@UnusedVariable # pylint: disable=unused-variable"
    outstr += nlind + "newline = '\\n'  # noqa: F841 #@UnusedVariable  # pylint: disable=unused-variable"
    outstr += nlind + "strlist.append('<head><style> td, th {border: 1px solid #ddd;  padding: 6px;} th {padding-top: 6px;padding-bottom: 6px;text-align: left;background-color: #0C6AA6; color: white;} </style></head> <body>')"

    # Open the workbook
    wb = load_workbook(filename = filename_in, read_only=False)

    # iterate through the sheets
    for ws in wb.worksheets:

        print(ws.title, '  ', ws.max_row, 'rows  ', ws.max_column, 'columns')

        # hack to work around when openpyxl ws.max_row returns more rows than actual
        num_rows:int = ws.max_row
        for row in ws.iter_rows(min_row=1, max_row=num_rows):
            if all(c.value is None for c in row):
                if row[0].row == num_rows + 2:  # two empty rows signifies the end of the help sheet
                    break
            else:
                num_rows = row[0].row

        # create the table name for the sheet name
        tbl_name = 'tbl_' + re.sub(r'[^A-Za-z0-9]','',ws.title)

        # build a formatted list of all rows in the sheet
        all_rows = generateRows(ws)

        # build a title string from the first row of the sheet
        titlestr = getTitle(all_rows,ws, wb.index(ws))

        # build a top notes table if there are any topnotes
        tbl_topnotes, nrows_topnotes = getNotes(all_rows,num_rows,tbl_name,notetype='top')

        # build a bottom notes table if there are any bottomnotes
        tbl_bottomnotes, nrows_bottomnotes = getNotes(all_rows,num_rows,tbl_name,notetype='bottom')

        # build the table components for the header row and the data rows
        if len(all_rows) > 1 + nrows_topnotes:
            tbl_fieldnames = getFieldnames(all_rows[1+nrows_topnotes],tbl_name)
            tbl_addrows = getAddrows(all_rows[2+nrows_topnotes:num_rows-nrows_bottomnotes],tbl_name)
        else:
            tbl_fieldnames = ''
            tbl_addrows = ''

        # assemble the py code in outstr
        # add the title
        outstr += titlestr

        # add  top notes
        if nrows_topnotes > 0:
            outstr += tbl_topnotes
            outstr += nlind + 'strlist.append(' + tbl_name + 'top' + '.get_html_string(attributes={' + note_table_attributes + '}))'

        # add the data table with header
        if len(tbl_addrows) > 0:
            outstr += nlind + tbl_name + ' = prettytable.PrettyTable()'
            outstr += nlind + tbl_fieldnames
            outstr += nlind + tbl_addrows
            outstr += nlind + 'strlist.append(' +  tbl_name + '.get_html_string(attributes={' + data_table_attributes + '}))'

        # add bottom notes
        if nrows_bottomnotes > 0:
            outstr += tbl_bottomnotes
            outstr += nlind + 'strlist.append(' + tbl_name + 'bottom' + '.get_html_string(attributes={' + note_table_attributes + '}))'

    # finalize outstr - py code
    outstr += nlind + "strlist.append('</body>')"

    outstr += nlind + "helpstr = ''.join(strlist)"

    # clean any html entities that get escaped by PrettyTable in its html output
    outstr += nlind + "return re.sub(r'&amp;', r'&',helpstr)" + '\n'

    return outstr

def writepyFile(filename_in:str, filename_out:str) -> None:
    outstr = buildpyCode(filename_in)

    # write outstr (py code) to the specified filename
    with open(filename_out, 'w', encoding='utf-8', newline='\n') as file_object:
        file_object.write(outstr)
    sleep(0.01)  #allow the previous write to settle, resolves appveyor file read fail

def writehtmlFile(_fname_in:str, filename_out:str, filename_htm:str) -> None:
    del _fname_in
    importfile = splitext(split(filename_out)[1])[0]
    importpath = abspath(split(filename_out)[0])
    sys.path.append(importpath)
    var = importlib.import_module(importfile)

    htmstr = var.content() + '\n'

    # write htmlstr (html) to the specified filename
    with open(filename_htm, 'w', encoding='utf-8', newline='\n') as file_object:
        file_object.write(htmstr)

if __name__ == '__main__':
    if len(sys.argv) > 1:
        currPath = (dirname(__file__)) + '/'
        #print(f"{currPath=}")
        if sys.argv[1] == 'all':
            #for filename in os.listdir(currPath + '../Input_files/'):
            for filename in listdir(currPath + '../Input_files/'):
                if filename.endswith('.xlsx'):
                    fn = filename.replace('.xlsx','')
                    fname_in =  currPath + '../Input_files/' + filename
                    fname_out = currPath + '../../../src/help/' + fn + '_help.py'
                    fname_htm = currPath + '../Output_html/' + fn + '_help.html'
                    print(f'\n{filename}')
                    writepyFile(fname_in,fname_out)
                    writehtmlFile(fname_in,fname_out,fname_htm)
        else:   #only one file
            fname_in =  currPath + '../Input_files/' + sys.argv[1] + '.xlsx'
            fname_out = currPath + '../../../src/help/' + sys.argv[1] + '_help.py'
            fname_htm = currPath + '../Output_html/' + sys.argv[1] + '_help.html'
            print(f'\n{sys.argv[1]}.xslx')
            writepyFile(fname_in,fname_out)
            writehtmlFile(fname_in,fname_out,fname_htm)
    else:
        print("Requires a <filename> to convert or 'all'")
        sys.exit()
