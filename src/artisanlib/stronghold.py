#
# ABOUT
# Stronghold XLS Roast Profile importer for Artisan

import openpyxl
import logging
from pathlib import Path
from typing import Final, Union, List, Dict, Tuple, Optional, Callable

from artisanlib.util import stringtoseconds, encodeLocalStrict
from artisanlib.atypes import ProfileData


_log: Final[logging.Logger] = logging.getLogger(__name__)

# returns a dict containing all profile information contained in the given Stronghold XLSX file
def extractProfileStrongholdXLSX(file:str,
        _etypesdefault:List[str],
        alt_etypesdefault:List[str],
        _artisanflavordefaultlabels:List[str],
        eventsExternal2InternalValue:Callable[[int],float]) -> ProfileData:

    res:ProfileData = ProfileData() # the interpreted data set

    res['roastertype'] = 'Stronghold'
    res['roasterheating'] = 3 # electric

    book = openpyxl.load_workbook(file)
    sheet = book.worksheets[0] # first sheet
    first_row = sheet[1]  # pyrefly: ignore[bad-argument-type]
    keys:Optional[List[str]] = None # data of columns with empty string as key are ignored
    machine:str = ''
    machine_size:float = 0
    if len(first_row) == 9: # S7Pro
        machine = 'S7 Pro'
        machine_size = 0.85
        # NOTE: column heads might be translated or edited
        # "Time" "Internal" "Tower Drum" "Hot Air Temp" "RoR" "Hot Air" "Halogen "Bean Agitation" "Note"
        keys = ['Time', 'ET', 'BT', 'IT', '', 'Air', 'Halogen', 'DrumSpeed', 'Note']

    elif len(first_row) == 10: # S2
        machine = 'S2'
        machine_size = 0.25
        # "Time" "Internal" "Bean Surface" "Hot Air Temp" "RoR" "BS RoR" "Hot Air" "Halogen" "Bean Agitation" "Note"
        keys = ['Time', 'ET', 'BT', 'IT', '', '', 'Air', 'Halogen', 'DrumSpeed', 'Note']

    elif len(first_row) == 12: # S7ProX
        machine = 'S7 ProX'
        machine_size = 0.85
        # "Time" "Internal" "Bean Surface" "Drum Surface" "Hot Air Temp" "RoR"  "B.S RoR"   "Hot Air" "Halogen" "Drum Heater" "Bean Agitation" "Note"
        keys = ['Time', 'ET', 'BT', 'DT', 'IT', '', '', 'Air', 'Halogen', 'DrumHeater', 'DrumSpeed', 'Note']

    elif len(first_row) == 13: # S8X/S9X
        machine = 'S8X/S9X'
#        machine_size = 4.5 or 8
        # "Time" "Internal" "Bean Surface" "Drum Surface" "Hot Air Temp" "RoR" "B.S RoR" "Hot Air" "Halogen" "Drum Heater" "Bean Agitation" "Blower" "Note"
        keys = ['Time', 'ET', 'BT', 'DT', 'IT', '', '', 'Air', 'Halogen', 'DrumHeater', 'DrumSpeed', 'Blower', 'Note']

    if keys is not None and sheet.max_row is not None:
        # import
        data:Dict[str, List[Union[float,int,str]]] = {}
        # read keys
        for key in keys:
            if key != '':
                data[key] = []
        # extract data
        for row in range(2, sheet.max_row + 1): # pyrefly: ignore[bad-argument-type]
            try:
                for i, key in enumerate(keys):
                    if key != '':
                        if key == 'Time':
                            data[key].append(stringtoseconds(str(sheet.cell(row, i+1).value))) # pyrefly: ignore[bad-argument-type]
                        else:
                            data[key].append(sheet.cell(row, i+1).value) # type:ignore
            except Exception as e:  # pylint: disable=broad-except
                _log.error(e)

        # convert data
        if 'Time' in data:
            res['title'] = encodeLocalStrict(Path(file).stem)
            res['roastertype'] = machine
            res['roastersize'] = machine_size
            res['timex'] = data['Time'] # type:ignore
            tx_len = len(res['timex'])
            res['mode'] = 'C'
            # add CHARGE/DROP
            res['timeindex'] = [0,0,0,0,0,0,max(0,tx_len-1),0]
            # add FCs
            if 'Note' in data and len(data['Note']) == tx_len:
                try:
                    res['timeindex'][2] = data['Note'].index('1st')
                except Exception:  # pylint: disable=broad-except
                    pass
            # add ET/BT
            if 'ET' in data and len(data['ET']) == tx_len:
                res['temp1'] = data['ET'] # type:ignore
            else:
                res['temp2'] = [-1.0]*tx_len
            if 'BT' in data and len(data['BT']) == tx_len:
                res['temp2'] = data['BT'] # type:ignore
            else:
                res['temp2'] = [-1.0]*tx_len
            res['extradevices'] = [25] # one extra virtual device
            res['extraname1'] = ['IT']
            res['extraname2'] = ['DT']
            res['extratimex'] = [res['timex'][:]] # pyright:ignore
            if 'IT' in data and len(data['IT']) == tx_len:
                res['extratemp1'] = [data['IT']] # type:ignore
            else:
                res['extratemp1'] = [[-1.0]*tx_len]
            if 'DT' in data and len(data['DT']) == tx_len:
                res['extratemp2'] = [data['DT']] # type:ignore
            else:
                res['extratemp2'] = [[-1.0]*tx_len]
            res['extramathexpression1'] = ['']
            res['extramathexpression2'] = ['']
            res['extraNoneTempHint1'] = [False] # apply temperature conversion
            res['extraNoneTempHint2'] = [False] # apply temperature conversion
            res['extraLCDvisibility1'] = [True]
            res['extraLCDvisibility2'] = [True]
            res['extraCurveVisibility1'] = [True]
            res['extraCurveVisibility2'] = [True]
            res['extraDelta1'] = [False]
            res['extraDelta2'] = [False]

            # add events

            events:List[Tuple[int, int, float]] = [] # list of triples (tx_idx, event_nr, value)
            try:
                for event_name, event_nr in [('DrumHeater', 0), ('DrumSpeed', 1), ('Halogen', 2), ('Air', 3), ('Blower', 4)]:
                    if event_name in data and len(data[event_name]) == tx_len:
                        last_value:float = -1
                        for i, v in enumerate(data[event_name]):
                            if isinstance(v, (int, float)):
                                if v != last_value:
                                    events.append((i, event_nr, v*10.0))
                                last_value = float(v)
                events.sort(key = lambda e : e[0])

                res['specialevents'] = [e[0] for e in events]
                res['specialeventstype'] = [e[1] for e in events]
                res['specialeventsvalue'] = [eventsExternal2InternalValue(int(round(float(e[2])))) for e in events]
                res['specialeventsStrings'] = ['']*len(events)
                res['etypes'] = alt_etypesdefault
                res['etypes'][0] = res['etypes'][1] + 'H'
                res['etypes'][0] = res['etypes'][0] + 'S'

            except Exception as e:  # pylint: disable=broad-except
                _log.exception(e)

    return res
